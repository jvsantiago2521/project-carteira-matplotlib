import matplotlib.pyplot as plt

from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

#Carregando arquivo excel
workbook: Workbook = load_workbook(WORKBOOK_PATH)
#Nome da planilha
sheetname = "Nova Planilha"
#Selecionando planilha
worksheet = Worksheet = workbook[sheetname]

row: tuple[Cell]
for row in worksheet.iter_rows(min_row=2):
    for cell in row:
        print(cell.value, end='\t')
    print()
#Alterando valor das celulas. (ABC... -> Coluna / 123... -> Linha)
#worksheet['B3'].value = 10

workbook.save(WORKBOOK_PATH)